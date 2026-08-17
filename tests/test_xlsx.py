"""Tester for XLSX-rapporten (§5).

Formatet finnes fordi CSV tvang Excel til å gjette på skilletegn og
tegnkoding, og gjettene motsa hverandre. Testene her leser fila tilbake med
openpyxl — altså gjennom en XLSX-implementasjon som ikke er vår egen.
"""

from __future__ import annotations

from openpyxl import load_workbook

from tfm_sjekk.modell import Alvorlighet, Funn
from tfm_sjekk.rapport import skriv_xlsx

NORSK = "«++11508=4310.001.12-QLF005» følger ikke TFM-grammatikken. Blåbærsyltetøy."


def funn() -> list[Funn]:
    return [
        Funn(
            kontroll="K2",
            alvorlighet=Alvorlighet.FEIL,
            melding=NORSK,
            global_id="1hqA2bC3dE4fG5hI6jK7lM",
            ifc_klasse="IfcFlowTerminal",
            kildefil="demo-rie.ifc",
            verdi="++11508=4310.001.12-QLF005",
        ),
        Funn(
            kontroll="K7",
            alvorlighet=Alvorlighet.INFO,
            melding="1 system i TFM-mastera er ikke brukt i modellen: 3600.002.04.",
        ),
    ]


def ark(tmp_path, funnliste=None):
    sti = skriv_xlsx(funnliste if funnliste is not None else funn(), tmp_path / "funn.xlsx")
    return load_workbook(sti).active


def test_overskrifter_og_rader(tmp_path):
    a = ark(tmp_path)
    assert [c.value for c in a[1]][:3] == ["Kontroll", "Grad", "Melding"]
    assert a.max_row == 3  # overskrift + to funn


def test_norske_tegn_uten_gjetting(tmp_path):
    """Selve poenget: ingen tegnkoding å tolke feil."""
    a = ark(tmp_path)
    assert a.cell(row=2, column=3).value == NORSK


def test_alvorlighet_skrives_som_tekst(tmp_path):
    """StrEnum-en skal bli «feil», ikke «Alvorlighet.FEIL»."""
    a = ark(tmp_path)
    assert a.cell(row=2, column=2).value == "feil"
    assert a.cell(row=3, column=2).value == "info"


def test_funn_uten_objekt_gir_tomme_celler(tmp_path):
    a = ark(tmp_path)
    assert a.cell(row=3, column=4).value is None  # global_id


def test_overskriftsraden_er_frosset_og_filtrerbar(tmp_path):
    """Det første en BIM-koordinator gjør er å filtrere på kontroll eller fag."""
    a = ark(tmp_path)
    assert a.freeze_panes == "A2"
    assert a.auto_filter.ref == "A1:G3"


def test_meldingskolonnen_brytes(tmp_path):
    """Norske setninger er lange; uten bryting blir arket ubrukelig bredt."""
    a = ark(tmp_path)
    assert a.cell(row=2, column=3).alignment.wrap_text is True


def test_tom_rapport_gir_gyldig_arbeidsbok(tmp_path):
    a = ark(tmp_path, [])
    assert a.max_row == 1
    assert a.auto_filter.ref is None  # filter over null rader gir korrupt fil
