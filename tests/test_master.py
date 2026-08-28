"""Tester for innlesing av TFM-master (§3, K7).

Mastera er det eneste inndataformatet som ikke er standardisert — den er et
regneark et menneske har laget. Testene her handler derfor mest om
slurvetoleranse: prefikser, hardt mellomrom, overskrifter som ikke står i
rad 1, ark med noe helt annet i.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from tfm_sjekk.config import MasterOppsett
from tfm_sjekk.feil import FilFeil
from tfm_sjekk.tabeller import les_master, normaliser


def skriv_csv(tmp_path, innhold: str, navn: str = "master.csv"):
    sti = tmp_path / navn
    sti.write_text(innhold, encoding="utf-8")
    return sti


def test_leser_semikolonseparert_csv(tmp_path):
    sti = skriv_csv(
        tmp_path,
        "systemforekomst;komponenttype\n3600.001.04;JVZ.001.008\n4310.001.12;QLF.001.001\n",
    )
    master = les_master(sti)
    assert master.systemer == {"3600.001.04", "4310.001.12"}
    assert master.komponenttyper == {"JVZ.001.008", "QLF.001.001"}
    assert master.kilde == "master.csv"


def test_leser_kommaseparert_csv(tmp_path):
    sti = skriv_csv(tmp_path, "system,komponenttype\n3600.001.04,JVZ.001.008\n")
    assert les_master(sti).systemer == {"3600.001.04"}


def test_normaliserer_prefikser_folk_skriver_inn(tmp_path):
    """«=3600.001.04» og «++115080=3600.001.04» er samme system."""
    sti = skriv_csv(
        tmp_path,
        "system\n=3600.001.04\n++115080=4310.001.12\n%JVZ.001.008\n",
    )
    assert les_master(sti).systemer == {"3600.001.04", "4310.001.12", "JVZ.001.008"}


def test_normaliser_takler_hardt_mellomrom_og_hermetegn():
    hardt = chr(0xA0)  # det Excel setter inn, og som str.strip() ikke fjerner
    assert normaliser(f"{hardt} 3600.001.04 {hardt}") == "3600.001.04"
    assert normaliser("'3600.001.04'") == "3600.001.04"
    assert normaliser("jvz.001.008") == "JVZ.001.008"


def test_finner_overskrift_under_tittelrader(tmp_path):
    """Ekte mastere har prosjektnavn og revisjonsrad over tabellen."""
    sti = skriv_csv(
        tmp_path,
        "Prosjekt 115080 Fiktivt bygg;;\nRevisjon C;2026-01-01;\n;;\n"
        "systemforekomst;komponenttype;beskrivelse\n"
        "3600.001.04;JVZ.001.008;Fiktivt luftbehandlingssystem\n",
    )
    master = les_master(sti)
    assert master.systemer == {"3600.001.04"}
    assert master.komponenttyper == {"JVZ.001.008"}


def test_feiler_hoylytt_uten_gjenkjennelig_kolonne(tmp_path):
    """En tom master ville fått K7 til å flagge hele modellen."""
    sti = skriv_csv(tmp_path, "kolonneA;kolonneB\nnoe;annet\n")
    with pytest.raises(FilFeil, match="gjenkjennelig kolonneoverskrift"):
        les_master(sti)


def test_feiler_pa_overskrift_uten_verdier(tmp_path):
    sti = skriv_csv(tmp_path, "systemforekomst;komponenttype\n;\n")
    with pytest.raises(FilFeil, match="ingen verdier"):
        les_master(sti)


def test_kolonnenavn_kan_konfigureres(tmp_path):
    sti = skriv_csv(tmp_path, "anlegg;\n3600.001.04;\n")
    with pytest.raises(FilFeil):
        les_master(sti)
    oppsett = MasterOppsett(kolonne_system=["anlegg"])
    assert les_master(sti, oppsett).systemer == {"3600.001.04"}


def test_leser_alle_ark_i_xlsx_og_hopper_over_forsiden(tmp_path):
    bok = Workbook()
    forside = bok.active
    forside.title = "Forside"
    forside.append(["TFM-master for prosjekt 115080"])
    forside.append(["Utarbeidet av RIE"])

    systemer = bok.create_sheet("Ark1")  # arknavnet skal ikke bety noe
    systemer.append(["Systemforekomst", "Beskrivelse"])
    systemer.append(["3600.001.04", "Fiktivt luftbehandlingssystem"])

    typer = bok.create_sheet("Typeliste rev C")
    typer.append(["Komponenttype"])
    typer.append(["JVZ.001.008"])

    sti = tmp_path / "master.xlsx"
    bok.save(sti)

    master = les_master(sti)
    assert master.systemer == {"3600.001.04"}
    assert master.komponenttyper == {"JVZ.001.008"}


def test_xlsx_tall_blir_ikke_flyttall(tmp_path):
    """«115080» skrevet inn som tall skal ikke bli «115080.0»."""
    bok = Workbook()
    ark = bok.active
    ark.append(["Systemforekomst"])
    ark.append([3600])
    sti = tmp_path / "tall.xlsx"
    bok.save(sti)

    assert les_master(sti).systemer == {"3600"}
