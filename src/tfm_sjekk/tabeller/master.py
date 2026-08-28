"""Prosjektets TFM-master (§3, K7).

SIMBA krever at prosjektet utarbeider en prosjektspesifikk TFM-master med
tverrfaglig systemliste, komponentliste, komponentforekomster og
komponenttyper. K7 sjekker modellen mot denne — begge veier.

Innlesingen er bevisst tolerant. Mastera er et regneark et menneske har
vedlikeholdt: den har logorad og revisjonstabell over overskriftene, den har
arknavn ingen kan gjette («Ark1», «Systemliste rev C»), og verdiene er
skrevet inn med de prefiksene folk husker fra TFM-ID-en («=3600.001.04»,
«++115080=3600.001.04»). Alt dette normaliseres bort her, slik at
kontrollen kan sammenligne rene nøkler.

Det som *ikke* gjettes er hvilke kolonner som betyr hva — de kommer fra
`tfm-sjekk.toml` (§14). Uten en gjenkjennelig kolonne feiler innlesingen
høylytt: en tom master ville fått K7 til å flagge hvert eneste objekt i
modellen, og det er en verre feil enn å stoppe.
"""

from __future__ import annotations

import csv
from collections.abc import Iterable, Iterator
from pathlib import Path

from pydantic import BaseModel

from tfm_sjekk.config import MasterOppsett
from tfm_sjekk.feil import FilFeil

# Excel og Word setter dette inn av seg selv; det ser ut som mellomrom,
# men str.strip() rører det ikke.
HARDT_MELLOMROM = chr(0xA0)


def normaliser(verdi: str) -> str:
    """Gjør en celle sammenlignbar med en parset TFM-ID.

    ``«++115080=3600.001.04»`` → ``3600.001.04``. Både mastera og modellen
    normaliseres med denne, så det holder at de to er enige om innholdet —
    ikke om skrivemåten.
    """
    tekst = verdi.replace(HARDT_MELLOMROM, " ").strip().strip("'\"").upper()
    if "=" in tekst:
        # ++115080=3600.001.04 → 3600.001.04. Plasseringen hører til bygget,
        # ikke systemet, og mastera er per prosjekt.
        tekst = tekst.rsplit("=", 1)[1]
    return tekst.lstrip("+-%").strip()


class TfmMaster(BaseModel):
    """Systemer og komponenttyper prosjektet har definert."""

    kilde: str
    systemer: set[str] = set()
    komponenttyper: set[str] = set()
    komponentforekomster: set[str] = set()

    @property
    def tom(self) -> bool:
        return not (self.systemer or self.komponenttyper or self.komponentforekomster)

    def kjenner_system(self, systemforekomst: str) -> bool:
        return normaliser(systemforekomst) in self.systemer

    def kjenner_type(self, komponenttype: str) -> bool:
        return normaliser(komponenttype) in self.komponenttyper


def les_master(sti: Path, oppsett: MasterOppsett | None = None) -> TfmMaster:
    """Leser TFM-master fra XLSX eller CSV.

    XLSX leses ark for ark; CSV er ett «ark». Kolonnene finnes ved navn, se
    `MasterOppsett`.
    """
    oppsett = oppsett or MasterOppsett()
    regneark = sti.suffix.lower() in (".xlsx", ".xlsm")
    ark = _les_xlsx(sti) if regneark else _les_csv(sti)

    master = TfmMaster(kilde=sti.name)
    gjenkjente_ark = 0
    try:
        for rader in ark:
            if _les_ark(rader, oppsett, master):
                gjenkjente_ark += 1
    except FilFeil:
        raise
    except OSError as feil:
        raise FilFeil(sti, f"kunne ikke leses: {feil.strerror or feil}.") from feil
    except Exception as feil:
        # BadZipFile fra openpyxl gjennom seksti linjer zipfile er ikke en
        # melding. Den vanligste årsaken er at fila ikke er det endelsen lover
        # — en CSV eller en HTML-tabell døpt om til .xlsx.
        hva = "regneark" if regneark else "CSV"
        raise FilFeil(sti, f"lot seg ikke lese som {hva}: {feil}") from feil

    if gjenkjente_ark == 0:
        forventet = ", ".join(
            sorted(
                {
                    *oppsett.kolonne_system,
                    *oppsett.kolonne_komponenttype,
                    *oppsett.kolonne_komponentforekomst,
                }
            )
        )
        raise FilFeil(
            sti,
            f"har ingen gjenkjennelig kolonneoverskrift i de "
            f"{oppsett.maks_overskriftsrader} første radene. Forventet én av: "
            f"{forventet}. Sett kolonnenavnene under [master] i tfm-sjekk.toml.",
        )
    if master.tom:
        raise FilFeil(
            sti,
            "har kolonneoverskriftene, men ingen verdier under dem. En tom "
            "master ville fått K7 til å flagge hele modellen.",
        )
    return master


def _les_ark(rader: list[list[str]], oppsett: MasterOppsett, master: TfmMaster) -> bool:
    """Leser ett ark inn i `master`. Returnerer False hvis arket ikke kjennes igjen."""
    kolonner = _finn_kolonner(rader, oppsett)
    if kolonner is None:
        return False

    overskriftsrad, plassering = kolonner
    mengder = {indeks: getattr(master, felt) for indeks, felt in plassering.items()}
    for rad in rader[overskriftsrad + 1 :]:
        for indeks, mengde in mengder.items():
            if indeks >= len(rad):
                continue
            verdi = normaliser(rad[indeks])
            if verdi:
                mengde.add(verdi)
    return True


def _finn_kolonner(
    rader: list[list[str]], oppsett: MasterOppsett
) -> tuple[int, dict[int, str]] | None:
    """Finner overskriftsraden og hvilket felt i `TfmMaster` hver kolonne fyller.

    Returnerer None hvis ingen av de konfigurerte kolonnenavnene finnes —
    da er dette et ark med noe annet i (forside, revisjonslogg, forklaringer).
    """
    kategorier = {
        "systemer": oppsett.kolonne_system,
        "komponenttyper": oppsett.kolonne_komponenttype,
        "komponentforekomster": oppsett.kolonne_komponentforekomst,
    }

    for radnummer, rad in enumerate(rader[: oppsett.maks_overskriftsrader]):
        plassering: dict[int, str] = {}
        for indeks, celle in enumerate(rad):
            overskrift = celle.replace(HARDT_MELLOMROM, " ").strip().lower()
            if not overskrift:
                continue
            for felt, kandidater in kategorier.items():
                if overskrift in kandidater and indeks not in plassering:
                    plassering[indeks] = felt
                    break
        if plassering:
            return radnummer, plassering
    return None


def _les_csv(sti: Path) -> Iterator[list[list[str]]]:
    """CSV er ett ark. Semikolon *og* komma godtas — norske Excel-eksporter
    bruker semikolon, slik som i `les_kodetabell`."""
    tekst = sti.read_text(encoding="utf-8-sig")
    linjer = tekst.splitlines()
    if not linjer:
        raise FilFeil(sti, "er tom. TFM-mastera skal ha minst en overskriftsrad.")
    skilletegn = ";" if linjer[0].count(";") > linjer[0].count(",") else ","
    yield [list(rad) for rad in csv.reader(linjer, delimiter=skilletegn)]


def _les_xlsx(sti: Path) -> Iterator[list[list[str]]]:
    """Alle ark, som tekst.

    `data_only=True` gir siste beregnede verdi framfor formelen — mastere er
    fulle av sammensatte celler à la ``=B2&"."&C2``.
    """
    from openpyxl import load_workbook

    bok = load_workbook(sti, read_only=True, data_only=True)
    try:
        for ark in bok.worksheets:
            yield [_som_tekst(rad) for rad in ark.iter_rows(values_only=True)]
    finally:
        bok.close()


def _som_tekst(rad: Iterable[object]) -> list[str]:
    """Tall blir tekst uten `.0`-hale — «115080» skrevet som tall i Excel skal
    ikke bli «115080.0»."""
    celler = []
    for celle in rad:
        if celle is None:
            celler.append("")
        elif isinstance(celle, float) and celle.is_integer():
            celler.append(str(int(celle)))
        else:
            celler.append(str(celle))
    return celler
