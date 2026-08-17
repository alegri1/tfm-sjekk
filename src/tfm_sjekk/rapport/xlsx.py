"""XLSX-rapport — for videre analyse i Excel (§5).

CSV er et tekstformat med to ting Excel må gjette: skilletegnet og
tegnkodingen. Begge gjettene slår feil et sted. Skrivebords-Excel deler på
listeskilletegnet fra regionsinnstillingene, mens Excel på web antar komma;
og legger man inn en `sep=`-linje for å styre det, bytter Excel parse-vei og
slutter å bry seg om BOM-en — da blir «følger» til «fÃ¸lger».

En ekte .xlsx har ingenting å gjette. Tegn og kolonner er strukturert i
fila, og den åpner likt i begge utgavene av Excel.

`openpyxl` er allerede en avhengighet fordi TFM-mastera kan være XLSX, så
formatet koster oss ingenting ekstra. CSV-en blir liggende ved siden av som
det maskinlesbare alternativet.
"""

from __future__ import annotations

from pathlib import Path

from tfm_sjekk.modell import Funn
from tfm_sjekk.rapport.csv_rapport import KOLONNER

OVERSKRIFTER = {
    "kontroll": ("Kontroll", 10),
    "alvorlighet": ("Grad", 11),
    "melding": ("Melding", 80),
    "global_id": ("GlobalId", 24),
    "ifc_klasse": ("IFC-klasse", 22),
    "kildefil": ("Fil", 20),
    "verdi": ("TFM-verdi", 38),
}


def skriv_xlsx(funn: list[Funn], sti: Path) -> Path:
    """Skriver funnene som Excel-arbeidsbok."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sti.parent.mkdir(parents=True, exist_ok=True)

    bok = Workbook()
    ark = bok.active
    ark.title = "Funn"

    ark.append([OVERSKRIFTER[k][0] for k in KOLONNER])
    for f in funn:
        rad = f.model_dump(include=set(KOLONNER))
        rad["alvorlighet"] = f.alvorlighet.value
        ark.append([rad[k] for k in KOLONNER])

    fet = Font(bold=True)
    fyll = PatternFill("solid", start_color="FFF4F4F4")
    for celle in ark[1]:
        celle.font = fet
        celle.fill = fyll

    for nummer, kolonne in enumerate(KOLONNER, start=1):
        ark.column_dimensions[get_column_letter(nummer)].width = OVERSKRIFTER[kolonne][1]

    # Meldingene er lange norske setninger; uten bryting blir arket ubrukelig bredt.
    meldingskolonne = get_column_letter(KOLONNER.index("melding") + 1)
    for celle in ark[meldingskolonne]:
        celle.alignment = Alignment(wrap_text=True, vertical="top")

    # Frys overskriftsraden og slå på filter: det første en BIM-koordinator
    # gjør er å filtrere på kontroll eller fag.
    ark.freeze_panes = "A2"
    if funn:
        ark.auto_filter.ref = f"A1:{get_column_letter(len(KOLONNER))}{len(funn) + 1}"

    bok.save(sti)
    return sti
